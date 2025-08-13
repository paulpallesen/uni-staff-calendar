#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook
from datetime import datetime, date, time, timedelta, UTC
import hashlib, os, sys

DEFAULT_TZ = "Australia/Sydney"

AUS_TZ_VTIMEZONE = """BEGIN:VTIMEZONE
TZID:Australia/Sydney
BEGIN:STANDARD
DTSTART:19700405T030000
TZOFFSETFROM:+1100
TZOFFSETTO:+1000
TZNAME:AEST
RRULE:FREQ=YEARLY;BYMONTH=4;BYDAY=1SU
END:STANDARD
BEGIN:DAYLIGHT
DTSTART:19701004T020000
TZOFFSETFROM:+1000
TZOFFSETTO:+1100
TZNAME:AEDT
RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=1SU
END:DAYLIGHT
END:VTIMEZONE
"""

def norm(s):  # normalize header names
    return (s or "").strip().lower()

def fmt_local(dt: datetime) -> str:
    return dt.strftime("%Y%m%dT%H%M%S")

def fmt_date(d: date) -> str:
    return d.strftime("%Y%m%d")

def _to_date(val):
    if val is None or val == "": return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    try: return datetime.fromisoformat(s).date()
    except Exception: return None

def _to_time(val):
    if val in (None, ""): return None
    if isinstance(val, datetime): return val.time().replace(second=0, microsecond=0)
    if isinstance(val, time): return val.replace(second=0, microsecond=0)
    s = str(val).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try: return datetime.strptime(s, fmt).time()
        except ValueError: pass
    try: return datetime.fromisoformat(s).time().replace(second=0, microsecond=0)
    except Exception: return None

def parse_datetime(d_val, t_val):
    d = _to_date(d_val)
    if not d: return None
    t = _to_time(t_val) or time(0,0,0)
    return datetime(d.year, d.month, d.day, t.hour, t.minute, 0)

def truthy(val) -> bool:
    if val is None: return False
    return str(val).strip().lower() in {"true","yes","y","1","transparent","free"}

def make_uid(fields):
    h = hashlib.sha1("|".join(str(x) for x in fields).encode("utf-8")).hexdigest()[:16]
    return f"{h}@youruni"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default=None)   # auto-use first sheet if not provided
    ap.add_argument("--out", default="calendar.ics")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    # Read & normalize headers
    hdr_raw = [ (h or "") for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) ]
    hdr_norm = [ norm(h) for h in hdr_raw ]
    hmap = { hdr_norm[i]: i for i in range(len(hdr_norm)) }

    if args.verbose:
        print("Detected headers (raw):", hdr_raw)
        print("Detected headers (normalized):", hdr_norm)

    def col(*aliases):
        # find the first header that exists among provided aliases
        for a in aliases:
            j = hmap.get(norm(a), -1)
            if j >= 0: return j
        return -1

    c_UID   = col("unique id","uid")
    c_Crs   = col("course code","course","code")
    c_Title = col("title","event title","headline","event headline")
    c_Cat   = col("category")
    c_SDate = col("start date","start")
    c_STime = col("start time")
    c_EDate = col("end date")
    c_ETime = col("end time")
    c_TZ    = col("timezone","tz")
    c_Loc   = col("location","place")
    c_Desc  = col("description","notes","note")
    c_URL   = col("link","url")
    c_TRAN  = col("transparent","transp")

    now_utc = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")

    lines = [
        "BEGIN:VCALENDAR",
        "PRODID:-//YourUni//Class Feeds 1.0//EN",
        "VERSION:2.0",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        AUS_TZ_VTIMEZONE.strip()
    ]

    event_count = 0
    rownum = 1

    for r in ws.iter_rows(min_row=2, values_only=True):
        rownum += 1
        if r is None or all(v in (None, "") for v in r):
            if args.verbose: print(f"Row {rownum}: empty, skipped")
            continue

        title_raw = (r[c_Title] if c_Title >= 0 else "")
        title = (str(title_raw).strip() if title_raw is not None else "")
        if args.verbose: print(f"Row {rownum}: Title raw -> {repr(title_raw)}")

        if not title:
            if args.verbose: print(f"Row {rownum}: no Title, skipped")
            continue

        sdate = r[c_SDate] if c_SDate >= 0 else None
        if not _to_date(sdate):
            if args.verbose: print(f"Row {rownum}: bad/missing Start Date -> {repr(sdate)}, skipped")
            continue

        uid  = (str(r[c_UID]).strip() if c_UID>=0 and r[c_UID] is not None else "")
        course = (str(r[c_Crs]).strip() if c_Crs>=0 and r[c_Crs] is not None else "")
        cat  = (str(r[c_Cat]).strip() if c_Cat>=0 and r[c_Cat] is not None else "")
        stime = r[c_STime] if c_STime>=0 else ""
        edate = r[c_EDate] if c_EDate>=0 else ""
        etime = r[c_ETime] if c_ETime>=0 else ""
        tz = (str(r[c_TZ]).strip() if c_TZ>=0 and r[c_TZ] is not None else DEFAULT_TZ)
        location = (str(r[c_Loc]).strip() if c_Loc>=0 and r[c_Loc] is not None else "")
        desc = (str(r[c_Desc]).strip() if c_Desc>=0 and r[c_Desc] is not None else "")
        url = (str(r[c_URL]).strip() if c_URL>=0 and r[c_URL] is not None else "")
        is_transparent = truthy(r[c_TRAN]) if c_TRAN>=0 else False

        is_all_day = (not _to_time(stime) and not _to_time(etime))
        if not uid:
            uid = make_uid([course, title, sdate, edate, stime, etime, location])

        summary = f"{course} — {title}" if course else title

        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}")
        lines.append(f"DTSTAMP:{now_utc}")
        lines.append(f"SUMMARY:{summary}")
        lines.append(f"TRANSP:{'TRANSPARENT' if is_transparent else 'OPAQUE'}")
        if location: lines.append(f"LOCATION:{location}")
        if desc: lines.append("DESCRIPTION:" + desc.replace("\\n","\\n"))
        if url: lines.append(f"URL:{url}")

        cats=[]
        if course: cats.append(course)
        if cat: cats.append(cat)
        if location: cats.append(location)
        if cats: lines.append(f"CATEGORIES:{','.join(cats)}")

        if is_all_day:
            start_d = _to_date(sdate)
            if edate:
                end_d = _to_date(edate) or start_d
            else:
                end_d = start_d
            end_excl = end_d + timedelta(days=1)
            lines.append(f"DTSTART;VALUE=DATE:{fmt_date(start_d)}")
            lines.append(f"DTEND;VALUE=DATE:{fmt_date(end_excl)}")
        else:
            dt_start = parse_datetime(sdate, stime)
            dt_end   = parse_datetime(edate or sdate, etime or stime or "00:00")
            if not dt_start or not dt_end:
                if args.verbose: print(f"Row {rownum}: bad time(s) -> start {repr(stime)} end {repr(etime)}, skipped")
                lines.append("END:VEVENT")
                continue
            lines.append(f"DTSTART;TZID={tz}:{fmt_local(dt_start)}")
            lines.append(f"DTEND;TZID={tz}:{fmt_local(dt_end)}")

        lines.append("END:VEVENT")
        event_count += 1

    out_dir = os.path.dirname(os.path.abspath(args.out))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    with open(args.out, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))

    print(f"Wrote {args.out} with {event_count} events")
    if event_count == 0:
        print("NOTE: 0 events written. Check that Title + Start Date have values, and dates are true Excel dates (not text).")

if __name__ == "__main__":
    main()

